import asyncio
import pandas as pd
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from src.model.question import Question
from src.settings.base import get_db  
import shutil
import os
from uuid import uuid4
from openpyxl import load_workbook

router = APIRouter()

UPLOAD_DIR = "uploads/"
os.makedirs(UPLOAD_DIR, exist_ok=True)

async def save_file_from_excel(image, row_idx, col_name) -> str:
    if image:
        filename = f"{uuid4()}_{row_idx}_{col_name}.png"
        filepath = os.path.join(UPLOAD_DIR, filename)
        image.save(filepath)
        return filepath
    return None

@router.post("/upload")
async def upload_excel(
    file: UploadFile,
    
    db: AsyncSession = Depends(get_db),
):
    try:
        loop = asyncio.get_running_loop()
        file_location = f"{UPLOAD_DIR}/{uuid4()}_{file.filename}"

        with open(file_location, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        workbook = await loop.run_in_executor(None, load_workbook, file_location)
        sheet = workbook.active

        # Extract images
        images = {img.anchor._from.row: img for img in sheet._images}  # Row index as key

        uploaded_questions = []
        
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            text_or_image = None
            option_a = None
            option_b = None
            option_c = None
            option_d = None

            # Extract text or image for Question
            text_or_image = row[0] if isinstance(row[0], str) else await save_file_from_excel(images.get(idx), idx, "question")

            # Extract text or image for Options
            option_a = row[1] if isinstance(row[1], str) else await save_file_from_excel(images.get(idx + 1), idx, "A")
            option_b = row[2] if isinstance(row[2], str) else await save_file_from_excel(images.get(idx + 2), idx, "B")
            option_c = row[3] if isinstance(row[3], str) else await save_file_from_excel(images.get(idx + 3), idx, "C")
            option_d = row[4] if isinstance(row[4], str) else await save_file_from_excel(images.get(idx + 4), idx, "D")

            # Save to DB
            db_question = Question(
                text=text_or_image if not text_or_image.startswith("uploads/") else None,
                image=text_or_image if text_or_image.startswith("uploads/") else None,
                option_a=option_a if not option_a.startswith("uploads/") else None,
                option_a_image=option_a if option_a.startswith("uploads/") else None,
                option_b=option_b if not option_b.startswith("uploads/") else None,
                option_b_image=option_b if option_b.startswith("uploads/") else None,
                option_c=option_c if not option_c.startswith("uploads/") else None,
                option_c_image=option_c if option_c.startswith("uploads/") else None,
                option_d=option_d if not option_d.startswith("uploads/") else None,
                option_d_image=option_d if option_d.startswith("uploads/") else None,
            )
            db.add(db_question)

            uploaded_questions.append({
                "question_text": text_or_image if not text_or_image.startswith("uploads/") else None,
                "image": text_or_image if text_or_image.startswith("uploads/") else None,
                "option_a": option_a if not option_a.startswith("uploads/") else None,
                "option_a_image": option_a if option_a.startswith("uploads/") else None,
                "option_b": option_b if not option_b.startswith("uploads/") else None,
                "option_b_image": option_b if option_b.startswith("uploads/") else None,
                "option_c": option_c if not option_c.startswith("uploads/") else None,
                "option_c_image": option_c if option_c.startswith("uploads/") else None,
                "option_d": option_d if not option_d.startswith("uploads/") else None,
                "option_d_image": option_d if option_d.startswith("uploads/") else None,
            })

        await db.commit()
        return uploaded_questions

    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")
